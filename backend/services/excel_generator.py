"""
Excel Generator — WBS Dual-View Export
  Sheet 1 "WBS Detail"  : Feature header rows → subtask rows with Human & Agent hours
  Sheet 2 "Task Matrix" : 7 type columns (R&D | UI/UX | Frontend | DB | Backend | Unit Testing | QA Testing)
                          No Feature column — each column lists ALL subtasks of that type
                          across all features, each cell = subtask name + description + H/A hours
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict
from collections import defaultdict
import os
from datetime import datetime

# ── Type order & colors ───────────────────────────────────────────────────────
TYPE_ORDER  = ["R&D", "UI/UX", "Frontend", "DB", "Backend", "Unit Testing", "QA Testing"]
TYPE_COLORS = {
    "R&D":          "FFF2CC",
    "UI/UX":        "FCE4D6",
    "Frontend":     "DDEBF7",
    "DB":           "E2EFDA",
    "Backend":      "EAD1DC",
    "Unit Testing": "F4CCFF",
    "QA Testing":   "D9EAD3",
}
NAVY        = "1E3A5F"
FEATURE_BLU = "2D6A9F"
WHITE       = "FFFFFF"
GREY        = "F2F2F2"


def _b():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _f(hex_):
    return PatternFill(start_color=hex_, end_color=hex_, fill_type="solid")


class ExcelGenerator:
    def __init__(self):
        self.export_dir = "temp/exports"
        os.makedirs(self.export_dir, exist_ok=True)

    def generate_excel(self, project_name: str, tasks: List[Dict]) -> str:
        grouped = self._group(tasks)
        wb = Workbook()

        self._sheet_detail(wb.active, grouped)
        self._sheet_matrix(wb.create_sheet("Task Matrix"), tasks)

        safe  = project_name.replace(" ", "_").replace("/", "_")
        ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
        path  = os.path.join(self.export_dir, f"{safe}_WBS_{ts}.xlsx")
        wb.save(path)
        return path

    # ── Group helper ──────────────────────────────────────────────────────────
    def _group(self, tasks):
        order, groups = [], {}
        for task in tasks:
            fid   = task.get("parent_id") or task.get("id", "UNK")
            fname = task.get("name", "")
            if fid not in groups:
                order.append(fid)
                label = fname.split(" - ", 1)[1] if " - " in fname else fname
                groups[fid] = {"name": label, "types": defaultdict(list)}
            groups[fid]["types"][task.get("task_type", "Dev")].append(task)
        return [(fid, groups[fid]) for fid in order]

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 1 — Vertical detail (unchanged good layout)
    # ─────────────────────────────────────────────────────────────────────────
    def _sheet_detail(self, ws, grouped):
        ws.title = "WBS Detail"
        col_cfg = [
            ("Feature", 26), ("Subtask Name", 35), ("Type", 14),
            ("Human (Hrs)", 13), ("Agent (Hrs)", 13), ("Description", 48),
        ]
        for i, (_, w) in enumerate(col_cfg, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 28
        for i, (lbl, _) in enumerate(col_cfg, 1):
            c = ws.cell(row=1, column=i, value=lbl)
            c.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            c.fill      = _f(NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _b()

        row = 2
        for fid, fdata in grouped:
            ws.row_dimensions[row].height = 22
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(col_cfg))
            c = ws.cell(row=row, column=1, value=f"  {fdata['name']}")
            c.font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
            c.fill      = _f(FEATURE_BLU)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = _b()
            row += 1

            for ttype in TYPE_ORDER:
                for task in fdata["types"].get(ttype, []):
                    ws.row_dimensions[row].height = 18
                    fill  = _f(TYPE_COLORS.get(ttype, WHITE))
                    h_hrs = task.get("human_hours") or task.get("duration_hours", 0)
                    a_hrs = task.get("agent_hours", 0)
                    description = task.get("description", "")

                    for ci, v in enumerate(
                        ["", description, ttype, h_hrs, a_hrs, description], 1
                    ):
                        c = ws.cell(row=row, column=ci, value=v)
                        c.font   = Font(size=10, name="Calibri")
                        c.fill   = fill
                        c.border = _b()
                        c.alignment = (
                            Alignment(horizontal="center", vertical="center")
                            if ci in (3, 4, 5)
                            else Alignment(vertical="center", wrap_text=True)
                        )
                    row += 1

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 2 — Task Matrix
    #   • 7 columns: one per task type (R&D, UI/UX, Frontend, DB, Backend, UT, QA)
    #   • No Feature column on the left
    #   • Each column lists ALL subtasks of that type from every feature
    #   • Each cell = subtask full name + description + H hrs / A hrs
    #   • Rows are independent per column; empty cells fill gaps
    # ─────────────────────────────────────────────────────────────────────────
    def _sheet_matrix(self, ws, tasks: List[Dict]):
        ws.title = "Task Matrix"

        num_types = len(TYPE_ORDER)

        # Column widths — wider to fit multi-line content
        for i in range(1, num_types + 1):
            ws.column_dimensions[get_column_letter(i)].width = 30

        # ── Header row ─────────────────────────────────────────────────────
        ws.row_dimensions[1].height = 32
        for ci, ttype in enumerate(TYPE_ORDER, 1):
            c = ws.cell(row=1, column=ci, value=ttype)
            c.font      = Font(bold=True, color="1E1E1E", size=11, name="Calibri")
            c.fill      = _f(TYPE_COLORS.get(ttype, GREY))
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = _b()

        # ── Build per-type task lists ──────────────────────────────────────
        # Each type gets a list of all tasks across all features
        type_task_lists: Dict[str, List[Dict]] = {t: [] for t in TYPE_ORDER}
        for task in tasks:
            ttype = task.get("task_type", "Dev")
            if ttype in type_task_lists:
                type_task_lists[ttype].append(task)

        # ── Write rows ────────────────────────────────────────────────────
        # We go row by row; the longest column determines total rows
        max_tasks = max((len(v) for v in type_task_lists.values()), default=0)

        for row_idx in range(max_tasks):
            row = row_idx + 2           # data starts at row 2
            ws.row_dimensions[row].height = 60   # tall enough for 3 lines

            for ci, ttype in enumerate(TYPE_ORDER, 1):
                task_list = type_task_lists[ttype]
                fill = _f(TYPE_COLORS.get(ttype, WHITE))

                if row_idx < len(task_list):
                    task = task_list[row_idx]
                    h_hrs = task.get("human_hours") or task.get("duration_hours", 0)
                    a_hrs = task.get("agent_hours", 0)

                    # Full subtask name (keep the full "Type - Feature" format)
                    full_name   = task.get("name", "")
                    description = task.get("description", "")

                    # Cell content: name, description, then hours
                    cell_val = f"{full_name}\n{description}\nH: {h_hrs}h  |  A: {a_hrs}h"
                    c = ws.cell(row=row, column=ci, value=cell_val)
                    c.font      = Font(size=9, name="Calibri")
                else:
                    # Empty — no task of this type at this index
                    c = ws.cell(row=row, column=ci, value="")
                    c.font = Font(size=9, name="Calibri")

                c.fill      = fill
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                c.border    = _b()

        # ── Totals row ────────────────────────────────────────────────────
        total_row = max_tasks + 2
        ws.row_dimensions[total_row].height = 28

        for ci, ttype in enumerate(TYPE_ORDER, 1):
            tasks_of_type = type_task_lists[ttype]
            tot_h = round(sum(float(t.get("human_hours") or t.get("duration_hours", 0)) for t in tasks_of_type), 1)
            tot_a = round(sum(float(t.get("agent_hours", 0)) for t in tasks_of_type), 1)
            count = len(tasks_of_type)

            c = ws.cell(row=total_row, column=ci,
                        value=f"Total: {count} tasks\nH: {tot_h}h  |  A: {tot_a}h")
            c.font      = Font(bold=True, size=9, name="Calibri")
            c.fill      = _f(NAVY)
            c.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = _b()
